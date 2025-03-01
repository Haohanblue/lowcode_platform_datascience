from fastapi import FastAPI, File, UploadFile, HTTPException, Form
from fastapi.responses import FileResponse
from fastapi.staticfiles import StaticFiles
from fastapi.middleware.cors import CORSMiddleware
from sqlalchemy import create_engine
import pandas as pd
import os
import uvicorn
from fastapi.responses import JSONResponse
import requests
import json
from chatGLM import get_zhipu_response
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import asyncio
from io import BytesIO
from contextlib import asynccontextmanager
# 使用生命周期事件处理器
@asynccontextmanager
async def lifespan(app: FastAPI):
    # 在应用启动时运行
    task_executor_task = asyncio.create_task(task_executor())
    yield
    # 在应用关闭时运行
    task_executor_task.cancel()
    await task_executor_task
# 这里是原第二段代码的内容开始
cwd = os.getcwd() 
app = FastAPI(lifespan=lifespan)  # 保留第二段代码的app实例

# 添加 CORS 中间件 (来自第一段代码)
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],  # 根据需求修改
    allow_credentials=True,
    allow_methods=["GET","POST","OPTIONS","PUT","DELETE","PATCH"],
    allow_headers=["*"],
)

# 用于保存识别结果的目录
UPLOAD_DIR = "uploadImages"
if not os.path.exists(UPLOAD_DIR):
    os.makedirs(UPLOAD_DIR)

# 用于保存识别结果的目录
UPLOAD_IMG_DIR = "static"
if not os.path.exists(UPLOAD_IMG_DIR):
    os.makedirs(UPLOAD_IMG_DIR)


import httpx

async def async_fetch(url):
    async with httpx.AsyncClient() as client:
        response = await client.get(url)
        return response.content


## 写一个图床的接口，接受上传的图片，保存至本地static文件夹下，返回图片的URL(已经挂载了静态文件目录)
@app.post("/upload_image/")
async def upload_img(task_id: str = Form(...), seq_id: int = Form(...), file: UploadFile = File(...)):
    try:
        print("task_id:", task_id)
        print("file_name:", file.filename)
        # 保存上传的图片文件
        # 用于保存识别结果的目录
        img_path = f"static/{task_id}"
        if not os.path.exists(img_path):
            os.makedirs(img_path)
        image_path = os.path.join(img_path, file.filename)
        with open(image_path, "wb") as f:
            f.write(file.file.read())
        # 创建存储任务ID的JSON文件路径
        task_file_path = os.path.join(UPLOAD_DIR, f"{task_id}.json")
                # 如果该任务的JSON文件不存在，创建一个空的JSON文件
        if not os.path.exists(task_file_path):
            task_data = {}  # 如果文件不存在，初始化为空字典
        else:
            # 如果文件存在，读取现有数据
            with open(task_file_path, "r", encoding='utf-8') as f:
                task_data = json.load(f)
        # 创建或更新对应顺序id的识别结果
        url_link = f"http://127.0.0.1:8000/{task_id}/{file.filename}"
        task_data[str(seq_id)] = {"image_name": file.filename, "ocr_result": url_link}
        # 将更新后的内容写回文件
        with open(task_file_path, "w", encoding='utf-8') as f:
            json.dump(task_data, f, ensure_ascii=False, indent=4)  # 使用 'w' 模式覆盖原文件内容
        # 返回任务ID和本次OCR识别结果
        return JSONResponse(content={
            "task_id": task_id,
            "seq_id": seq_id,
            "img_url": url_link,
            "image_name": file.filename,
            "status": "success"
        })
    except Exception as e:
        return JSONResponse(content={"error": str(e)}, status_code=400)


@app.post("/upload/")
async def upload_image(task_id: str = Form(...), seq_id: int = Form(...), file: UploadFile = File(...)):
    try:
        print("task_id:", task_id)
        print("file_name:", file.filename)
        
        # 创建存储任务ID的JSON文件路径
        task_file_path = os.path.join(UPLOAD_DIR, f"{task_id}.json")
        
        # 如果该任务的JSON文件不存在，创建一个空的JSON文件
        if not os.path.exists(task_file_path):
            task_data = {}  # 如果文件不存在，初始化为空字典
        else:
            # 如果文件存在，读取现有数据
            with open(task_file_path, "r", encoding='utf-8') as f:
                task_data = json.load(f)
        
        # 读取上传的文件内容
        img_bytes = await file.read()
        print("img_bytes:", len(img_bytes))
        
        # 这里替换为你实际的OCR调用代码
        r = requests.post(
            'http://127.0.0.1:8501/ocr', files={'image': img_bytes},
        )
        ocr_out = r.json().get('results', [])
        print(ocr_out)

        # 合并OCR结果
        text = ""
        for i in ocr_out:
            text += i.get("text", "")
        
        # 创建或更新对应顺序id的识别结果
        task_data[str(seq_id)] = {"image_name": file.filename, "ocr_result": text}
        
        # 将更新后的内容写回文件
        with open(task_file_path, "w", encoding='utf-8') as f:
            json.dump(task_data, f, ensure_ascii=False, indent=4)  # 使用 'w' 模式覆盖原文件内容

        # 返回任务ID和本次OCR识别结果
        return JSONResponse(content={
            "task_id": task_id,
            "seq_id": seq_id,
            "ocr_result": text,
            "image_name": file.filename,
            "status": "success"
        })

    except Exception as e:
        return JSONResponse(content={"error": str(e)}, status_code=400)

@app.get("/download/{task_id}")
async def download_task_result(task_id: str):
    try:
        # 创建存储任务ID的JSON文件路径
        task_file_path = os.path.join(UPLOAD_DIR, f"{task_id}.json")
        
        if not os.path.exists(task_file_path):
            raise FileNotFoundError("任务文件不存在")

        # 读取任务结果的JSON文件
        with open(task_file_path, "r", encoding='utf-8') as f:
            task_data = json.load(f)

        # 创建一个新的Excel工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = "OCR结果"

        # 设置表头
        ws.append(["任务序号", "图片名称", "识别结果"])

        # 遍历任务数据并填充到Excel表格中
        for seq_id, result in task_data.items():
            ws.append([seq_id, result["image_name"], result["ocr_result"]])

        # 调整列宽
        for col in range(1, 4):
            max_length = 0
            column = get_column_letter(col)
            for row in ws.iter_rows():
                cell = row[col - 1]
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width

        # 保存为临时文件
        excel_file_path = os.path.join(UPLOAD_DIR, f"{task_id}.xlsx")
        wb.save(excel_file_path)

        # 返回下载链接
        return FileResponse(excel_file_path, media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', filename=f"{task_id}.xlsx")

    except Exception as e:
        return JSONResponse(content={"error": str(e)}, status_code=400)



@app.post("/get_classify_result/")
async def get_classify_result(req_data: dict):
    try:
        # 这里替换为你实际的分类调用代码
        model = "glm-4-air-0111"
        print(req_data)
        requirement_template = req_data["classifyRequirement"]

        # requirement_template = """
        # 你是一位专业的金融行为分析师。请仔细分析以下投资者的社交媒体发言，判断该发言是否表现出投资者倾向于追求高风险-高回报的"彩票式股票"投资特征。
        # 分析框架
        # 请逐步分析以下要素：
        # 1. 风险认知指标：
        # * 发言是否显示出高风险承受意愿（如，"赌"、"无所谓风险"）
        # * 是否提及"破釜沉舟"式投资决策（如，"生死一搏"、"不成功便成仁"）
        # * 是否表现出"全部押注"倾向（如，"孤注一掷"、"全仓买入"、"一把梭哈"）
        # * 是否体现非理性乐观（如，忽略风险，仅提到收益）
        # * 是否偏好高波动性股票或偏好跟风炒作（如，提到"妖股"、"大起大落"，或提到"大家都在买"、"热点"）
        # 2. 收益预期指标：
        # * 是否期待极高收益回报（如，提及"翻倍"、"一夜暴富"、“涨停”）
        # * 是否有不切实际的收益预期（如，对短期回报持有过度乐观预期）
        # * 是否频繁提及"暴富"概念（如，"暴利"、"财富自由"）
        # * 是否过分强调短期收益（如，"明天就能赚"、"短期暴涨"）
        # * 是否表现出急于求成心态（如，提到"急于翻身"、"抓住最后机会"）


        # 量化评分维度
        # 请分别对以下维度进行0-1分的评分：
        # - 风险倾向：[0-分]1（1分表示风险规避，1分表示极度风险追求）
        # - 收益预期：[0-1分]（1分表示理性预期，1分表示极度非理性预期）
        # - 若无法根据文本识别出上述某一维度，则该维度赋值为'null'

        # 关键词识别
        # 请特别关注以下类型词汇的出现：
        # 赌博类：梭哈、豪赌、赌一把、孤注一掷、下注、中奖、彩票
        # 暴利类：暴富、翻倍、一夜暴富、翻身、财富自由、风口、黑马、独角兽、涨停
        # 极端类：破釜沉舟、生死一搏、破罐破摔、刀口舔血、放手一搏、富贵险中求
        # 机会类：抄底、抓住机会、押宝
        # 网络俚语：一波肥、吃鸡、躺赢、单车变摩托

  
        # **重要提示**
        # 1. 避免过度推断，分析应基于文本明确表达。
        # 2. 注意区分正常投资讨论和彩票式投资倾向。
        # 3. 考虑发言的完整语境，避免脱离上下文判断。
        # """
        # req_data = {
        #     "task_id": 1,
        #     "total_num": 20,
        #     "offset": 0,
        #     "limit": 20,
        #     "classify_num": 2,
        #     "classify_content": ["风险倾向", "收益预期"],
        #     "question_list": [
        #         {
        #             "question_id": 1,
        #             "question_content": "雄起赛力斯，威武华为，赛力斯牛逼[赞]",
        #         },
        #         {
        #             "question_id": 2,
        #             "question_content": "我要翻倍，明天就能赚，一夜暴富",
        #         },
        #         {
        #             "question_id": 3,
        #             "question_content": "被动杀跌（实为诱空）无需恐慌",
        #         }
        #     ],
        # }
        data = req_data["question_list"]

        # 将data从列表转化为字符串，以便于传递给模型
        classify_num = req_data["classify_num"]
        classify_cotent = req_data["classify_content"]


        data_str = json.dumps(data)
        true_template = f"""
        你是一个专业的分类模型。你需要根据下面所给的分类要求，对以下文本进行分类。
        严格按照分类要求进行分类。
        分类内容是由列表中的对象，对象里question_content的文本组成的，每个元素都是一个字符串。
        请按照分类要求，对每个文本进行分类，将每一个类别归属的概率从0-1之间打分，输出连续的值，
        结果为每一个类别对应一个值的对象，注意各个类别之间没有关系，不需要让他们的和等于1，单独评价即可，
        可以理解为使用sigmoid函数转化为0-1上的值。
        请注意，分类结果的顺序必须与文本的顺序一致。
        请注意，分类结果必须是一个字符串键值对，且键必须是分类要求中的一个，值为对应的值。
        请注意，只输出json格式的分类结果，不要携带其他任何内容。
        分类数量：<{classify_num}>
        分类类别：<{classify_cotent}>
        分类要求：<{requirement_template}>
        分类内容：<{data_str}>
        以下是一个输出格式的示例：
        当分类数量为2，分类类别为["风险倾向", "收益预期"]时，输出格式如下：
        注意这只是一个示例，实际输出的内容可能会有所不同。
        {{
            "data":[
                {{
                    "id":1,
                    "result":{{
                        "风险倾向":0.849977,
                        "收益预期":0.257986
                    }}
                }},
                {{
                    "id":2,
                    "result":{{
                        "风险倾向":0.987751,
                        "收益预期":0.466871
                    }}
                }}
            ]
        }}"""
        response = await get_zhipu_response(true_template, model)

        ## 提取response字符串中的json格式内容,```json\n{...}\n```内的内容
        response = response.split("```json\n")[1].split("\n```")[0]
        # 将字符串转化为字典
        response = json.loads(response)
        print(response)
        return JSONResponse(content={"task_id": req_data["task_id"], "response": response})

    except Exception as e:
        return JSONResponse(content={"error": str(e)}, status_code=400)




@app.post("/get_ocr_result/")
async def get_ocr_result(req_data: dict):
    try:
        # 这里替换为你实际的分类调用代码
        print("请求体：",req_data)
        data = req_data["question_list"]
        # 读取上传的文件内容
        res_list = []
        for x in data:
            ## 读取图片url，获取图片内容
            url = x["question_content"]
            id = x["question_id"]
            print("url:", url)
            img_bytes  = await async_fetch(url)
            print("img_bytes:", len(img_bytes))
            # 这里替换为你实际的OCR调用代码
            r = await httpx.AsyncClient().post(
                'http://127.0.0.1:8501/ocr', files={'image': img_bytes},
            )
            print("r:", r.json())
            ocr_out = r.json().get('results', [])
            print("ocr结果",ocr_out)
            # 合并OCR结果
            text = ""
            for i in ocr_out:
                text += i.get("text", "")
            j = {
                "id": id,
                "result": {"OCR结果": text}
            }
            res_list.append(j)
        response = {
            "data": res_list
        }
        print(response)
        return JSONResponse(content={"task_id": req_data["task_id"], "response": response})

    except Exception as e:
        return JSONResponse(content={"error": str(e)}, status_code=400)

# 用于存储任务队列
task_queue = asyncio.Queue()



# 定义文件存储路径
BASE_DIR = "tasks"
INPUT_DIR = os.path.join(BASE_DIR, "input")
OUTPUT_DIR = os.path.join(BASE_DIR, "output")
DICT_DIR = os.path.join(BASE_DIR, "dict")

# 确保目录存在
os.makedirs(INPUT_DIR, exist_ok=True)
os.makedirs(OUTPUT_DIR, exist_ok=True)
os.makedirs(DICT_DIR, exist_ok=True)

# 任务处理函数
async def process_task(task_id: str, task_type: str, file_path: str, other_data: dict):
    print("开始处理任务", task_id)
    task_file = os.path.join(DICT_DIR, f"{task_id}.json")
    try:
        # 更新任务状态为进行中
        update_task_status(task_id, "进行中")

        # 读取 Excel 文件
        df = pd.read_excel(file_path)
        print(df)
        if task_type == "分类任务":
            new_list = other_data["classify_content"].split(",")
            new_list.insert(0, "分类问题")
            # 新的列名列表
            # 检查新列名是否多于现有列数
            if len(new_list) > len(df.columns):
                # 计算需要添加的空列数量
                additional_columns = len(new_list) - len(df.columns)
                print("需要添加的空列数量", additional_columns)
                # 添加空列
                for i in range(additional_columns):
                    df[f'空列{i + 1}'] = None  # 使用 None 创建空列
            df.columns = new_list  # 替换为你的新列名
            print(df)

        else:
            df.columns = ['图片URL', '识别结果']  # 替换为你的新列名
        print(df)
        # 模拟处理，逐行处理并更新进度
        total_rows = df.shape[0]
        for index, row in df.iterrows():
            print("index:", index)
            print("row:", type(row))
            print(row)
            print(f"处理任务 {task_id} 的第 {index + 1} 行")
            await asyncio.sleep(0)  # 让出控制权，避免阻塞事件循环
            df.iloc[index] = await handle_data(row,task_type,other_data)
            # 更新进度
            await asyncio.sleep(0)  # 让出控制权，避免阻塞事件循环
            progress = (index + 1) / total_rows
            update_task_progress(task_id, progress)
            await asyncio.sleep(0)  # 让出控制权，避免阻塞事件循环

        # 假设处理完成后，保存结果到输出文件
        output_file_path = os.path.join(OUTPUT_DIR, f"{task_id}.xlsx")
        df.to_excel(output_file_path, index=False)

        # 更新任务状态为已完成
        update_task_status(task_id, "已完成")
        print("任务处理完成", task_id)
    except Exception as e:
        update_task_status(task_id, "失败", str(e))

async def handle_data(row,task_type,other_data):
    if task_type == "OCR任务":
        url_link = row["图片URL"]
        res = await get_ocr_formal_result(url_link)
        row["识别结果"] = res
    elif task_type == "分类任务":
        question = row["分类问题"]
        print("question:", question)
        res_obj = await get_classify_formal_result(question,other_data)
        answer = res_obj["response"]["data"]["result"]
        print("answer:", answer)
        for k, v in answer.items():
            row[k] = v
    return row


async def get_ocr_formal_result(url_link: str):
    img_bytes  = await async_fetch(url_link)
    print("img_bytes:", len(img_bytes))
    # 这里替换为你实际的OCR调用代码
    r = await httpx.AsyncClient().post(
        'http://127.0.0.1:8501/ocr', files={'image': img_bytes},
    )
    print("r:", r.json())
    ocr_out = r.json().get('results', [])
    print("ocr结果",ocr_out)
    # 合并OCR结果
    text = ""
    for i in ocr_out:
        text += i.get("text", "")
    return text

async def get_classify_formal_result(question,req_data: dict):
    try:
        # 这里替换为你实际的分类调用代码
        model = "glm-4-air-0111"
        requirement_template = req_data["classifyRequirement"]

        # requirement_template = """
        # 你是一位专业的金融行为分析师。请仔细分析以下投资者的社交媒体发言，判断该发言是否表现出投资者倾向于追求高风险-高回报的"彩票式股票"投资特征。
        # 分析框架
        # 请逐步分析以下要素：
        # 1. 风险认知指标：
        # * 发言是否显示出高风险承受意愿（如，"赌"、"无所谓风险"）
        # * 是否提及"破釜沉舟"式投资决策（如，"生死一搏"、"不成功便成仁"）
        # * 是否表现出"全部押注"倾向（如，"孤注一掷"、"全仓买入"、"一把梭哈"）
        # * 是否体现非理性乐观（如，忽略风险，仅提到收益）
        # * 是否偏好高波动性股票或偏好跟风炒作（如，提到"妖股"、"大起大落"，或提到"大家都在买"、"热点"）
        # 2. 收益预期指标：
        # * 是否期待极高收益回报（如，提及"翻倍"、"一夜暴富"、“涨停”）
        # * 是否有不切实际的收益预期（如，对短期回报持有过度乐观预期）
        # * 是否频繁提及"暴富"概念（如，"暴利"、"财富自由"）
        # * 是否过分强调短期收益（如，"明天就能赚"、"短期暴涨"）
        # * 是否表现出急于求成心态（如，提到"急于翻身"、"抓住最后机会"）


        # 量化评分维度
        # 请分别对以下维度进行0-1分的评分：
        # - 风险倾向：[0-分]1（1分表示风险规避，1分表示极度风险追求）
        # - 收益预期：[0-1分]（1分表示理性预期，1分表示极度非理性预期）
        # - 若无法根据文本识别出上述某一维度，则该维度赋值为'null'

        # 关键词识别
        # 请特别关注以下类型词汇的出现：
        # 赌博类：梭哈、豪赌、赌一把、孤注一掷、下注、中奖、彩票
        # 暴利类：暴富、翻倍、一夜暴富、翻身、财富自由、风口、黑马、独角兽、涨停
        # 极端类：破釜沉舟、生死一搏、破罐破摔、刀口舔血、放手一搏、富贵险中求
        # 机会类：抄底、抓住机会、押宝
        # 网络俚语：一波肥、吃鸡、躺赢、单车变摩托

  
        # **重要提示**
        # 1. 避免过度推断，分析应基于文本明确表达。
        # 2. 注意区分正常投资讨论和彩票式投资倾向。
        # 3. 考虑发言的完整语境，避免脱离上下文判断。
        # """
        # req_data = {
        #     "task_id": 1,
        #     "total_num": 20,
        #     "offset": 0,
        #     "limit": 20,
        #     "classify_num": 2,
        #     "classify_content": ["风险倾向", "收益预期"],
        #     "question_list": [
        #         {
        #             "question_id": 1,
        #             "question_content": "雄起赛力斯，威武华为，赛力斯牛逼[赞]",
        #         },
        #         {
        #             "question_id": 2,
        #             "question_content": "我要翻倍，明天就能赚，一夜暴富",
        #         },
        #         {
        #             "question_id": 3,
        #             "question_content": "被动杀跌（实为诱空）无需恐慌",
        #         }
        #     ],
        # }
        data = question

        # 将data从列表转化为字符串，以便于传递给模型
        classify_num = req_data["classify_num"]
        classify_cotent = req_data["classify_content"]


        data_str = json.dumps(data,ensure_ascii=False, indent=4)
        true_template = f"""
        你是一个专业的分类模型。你需要根据下面所给的分类要求，对以下文本进行分类。
        严格按照分类要求进行分类。
        分类内容是由一个字符串组成的文本问题，请你针对该文本进行分类。
        请按照分类要求，对该文本进行分类，将每一个类别归属的概率从0-1之间打分，输出连续的值，
        结果为每一个类别对应一个值的对象，注意各个类别之间没有关系，不需要让他们的和等于1，单独评价即可，
        可以理解为使用sigmoid函数转化为0-1上的值。
        请注意，分类结果必须是一个字符串键值对，且键必须是分类要求中的一个，值为对应的值。
        请注意，只输出json格式的分类结果，不要携带其他任何内容。
        请注意，问题只有一个，结果也应该只有一个。
        分类数量：<{classify_num}>
        分类类别：<{classify_cotent}>
        分类要求：<{requirement_template}>
        分类内容：<{data_str}>
        以下是一个输出格式的示例：
        当分类数量为2，分类类别为["风险倾向", "收益预期"]时，输出格式如下：
        注意这只是一个示例，实际输出的内容可能会有所不同。
        输出结果必须有数值，如果无法识别，请填写None
        {{
            "data":{{
                    "result":
                        {{
                        "风险倾向":0.849977,
                        "收益预期":0.257986
                    }}
            }}
        }}"""
        print("最终版提示词：",true_template)
        response = await get_zhipu_response(true_template, model)

        ## 提取response字符串中的json格式内容,```json\n{...}\n```内的内容
        response = response.split("```json\n")[1].split("\n```")[0]
        # 将字符串转化为字典
        response = json.loads(response)
        print(response)
        return {"response": response}

    except Exception as e:
        return "error"



def update_task_status(task_id: str, status: str, error_message: str = ""):
    print(f"更新任务 {task_id} 状态为 {status}，错误信息：{error_message}")
    task_file = os.path.join(DICT_DIR, f"{task_id}.json")
    with open(task_file, "r",encoding='utf-8') as f:
        task_data = json.load(f)
    task_data["status"] = status
    if error_message:
        task_data["error"] = error_message
    with open(task_file, "w",encoding='utf-8') as f:
        json.dump(task_data, f,ensure_ascii=False, indent=4)

def update_task_progress(task_id: str, progress: float):
    print(f"更新任务 {task_id} 进度为 {progress}")
    task_file = os.path.join(DICT_DIR, f"{task_id}.json")
    with open(task_file, "r",encoding='utf-8') as f:
        task_data = json.load(f)
    task_data["progress"] = progress
    with open(task_file, "w",encoding='utf-8') as f:
        json.dump(task_data, f,ensure_ascii=False, indent=4)


# 异步任务执行器
async def task_executor():
    while True:
        task = await task_queue.get()
        await process_task(**task)
        task_queue.task_done()

@app.get("/download_done_task/{task_id}")
async def download_done_file(task_id: str):
    try:
        # 构建安全路径
        filename = f"{task_id}.xlsx"
        file_path = f"{OUTPUT_DIR}/{filename}" 
        # 转换为绝对路径并验证是否在允许目录内
        print("file_path:", file_path)
        # 检查文件是否存在
        # 发送文件
        return FileResponse(
            path=file_path,
            filename=filename,  # 下载时显示的文件名
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    except HTTPException as he:
        raise he
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"服务器错误: {str(e)}")
    

@app.post("/start-task")
async def start_task(task_id: str = Form(...), task_type: str = Form(...), template:str = Form(...),classify_num:str = Form(...),classify_content:str = Form(...),file: UploadFile = File(...)):
    try:
        # 保存输入文件
        print("task_id:", task_id)
        print("task_type:", task_type)
        print("classify_content",classify_content)
        print(type(classify_content))
        input_file_path = os.path.join(INPUT_DIR, f"{task_id}.xlsx")
        with open(input_file_path, "wb") as f:
            f.write(await file.read())
        if task_type == "分类任务":
            other_data = {
                "classifyRequirement": template,
                "classify_num": classify_num,
                "classify_content": classify_content
            }
        else:
            other_data = {}
        # 初始化任务记录
        task_data = {
            "task_id": task_id,
            "task_type": task_type,
            "status": "等待中",
            "progress": 0.0,
            "input_file": input_file_path,
            "output_file": os.path.join(OUTPUT_DIR, f"{task_id}.xlsx"),
            "other_data": other_data
        }
        task_file = os.path.join(DICT_DIR, f"{task_id}.json")
        with open(task_file, "w", encoding='utf-8') as f:
            json.dump(task_data, f, ensure_ascii=False, indent=4)
                # 将任务添加到队列
        
        await task_queue.put({
            "task_id": task_id,
            "task_type": task_type,
            "file_path": input_file_path,
            "other_data": other_data
        })

        return JSONResponse(content={"message": f"任务 {task_id} 创建成功"})
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/task-status/{task_id}")
async def get_task_status(task_id: str):
    try:
        task_file = os.path.join(DICT_DIR, f"{task_id}.json")
        if not os.path.exists(task_file):
            raise HTTPException(status_code=404, detail="任务未找到")

        with open(task_file, "r", encoding='utf-8') as f:
            task_data = json.load(f)
        
        return JSONResponse(content=task_data)
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))






# 挂载静态文件目录（来自第二段代码）
app.mount("/", StaticFiles(directory="static", html=True), name="static")

# 这里是第一段代码的内容开始，将其合并进来
# 注：移除第一段代码中的 app = FastAPI() 和其 own uvicorn.run
#    以及保持路由和逻辑不变
# 加载环境变量              
UPLOAD_FOLDER = "./uploads"
RESULT_FOLDER = "./results"

os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(RESULT_FOLDER, exist_ok=True)



def connect_mysql():
    """连接 MySQL 数据库"""
    engine = create_engine('mysql+pymysql://sa:Datateam.001@10.132.166.96:3306/rawdata')
    return engine

def query_data(sql):
    """查询数据"""
    engine = connect_mysql()
    try:
        print("Executing SQL:", sql)
        with engine.connect() as conn:
            data = pd.read_sql(sql, conn)
    except Exception as e:
        print(f"Error executing SQL: {e}")
        data = pd.DataFrame()
    finally:
        engine.dispose()
    return data

def define_sql(data):
    """生成 SQL 查询语句"""
    origin = data.columns[0]
    target = data.columns[1]
    origin_data = [f"'{item}'" for item in data[origin]]

    if origin == 'code' and target == 'phone':
        sql = f"SELECT `用户 Code`, `手机` FROM leads WHERE `用户 Code` IN ({', '.join(origin_data)}) ORDER BY FIELD(`用户 Code`, {', '.join(origin_data)})"
    elif origin == 'code' and target == 'name':
        sql = f"SELECT `用户 Code`, `姓名` FROM leads WHERE `用户 Code` IN ({', '.join(origin_data)}) ORDER BY FIELD(`用户 Code`, {', '.join(origin_data)})"
    elif origin == 'phone' and target == 'code':
        sql = f"SELECT `手机`, `用户 Code` FROM leads WHERE `手机` IN ({', '.join(origin_data)}) ORDER BY FIELD(`手机`, {', '.join(origin_data)})"
    elif origin == 'code' and target == 'intention':
        sql = f"SELECT `用户 Code`, `意向编号` FROM intention WHERE `用户 Code` IN ({', '.join(origin_data)}) ORDER BY FIELD(`用户 Code`, {', '.join(origin_data)})"
    elif origin == 'intention' and target == 'code':
        sql = f"SELECT `意向编号`, `用户 Code` FROM intention WHERE `意向编号` IN ({', '.join(origin_data)}) ORDER BY FIELD(`意向编号`, {', '.join(origin_data)})"
    elif origin == 'phone' and target == 'intention':
        sql = f"SELECT `用户手机`, `意向编号` FROM intention WHERE `用户手机` IN ({', '.join(origin_data)}) ORDER BY FIELD(`用户手机`, {', '.join(origin_data)})"
    elif origin == 'intention' and target == 'phone':
        sql = f"SELECT `意向编号`, `用户手机` FROM intention WHERE `意向编号` IN ({', '.join(origin_data)}) ORDER BY FIELD(`意向编号`, {', '.join(origin_data)})"
    else:
        raise ValueError("Unsupported column mapping")

    return sql



# 保留主程序运行部分为一个即可
if __name__ == '__main__':
    uvicorn.run(app="main:app", host="0.0.0.0", port=8000, reload=True)